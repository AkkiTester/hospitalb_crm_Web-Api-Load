import openpyxl
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)
def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value
def writeData(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value=data
    workbook.save(file)

def alldata(path1,sheetName):
    data=[]
    for r in range (2,getRowCount(path1,sheetName)):

        data2=[]
        data2.append(readData(path1,sheetName,r,1))
        data2.append(readData(path1,sheetName,r,2))
        data.append(data2)
        data2=[]
    return data